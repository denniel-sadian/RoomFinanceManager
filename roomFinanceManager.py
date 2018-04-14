#! python3
# Denniel Luis Saway Sadian
# https://denniel-sadian.github.io/
# February 18, 2018
"""
This program needs MS Excel for generating the report.
And I'm afraid that you'll face some issue if you'll
try to run it on any non-Windows machine.
"""

try:
    from tkinter import *
    from tkinter import ttk, messagebox
    from stringy import *
    from time import strftime
    from threading import Thread
    from about_dialog import AboutDialog
    from openpyxl.styles import Alignment, PatternFill, Color
    import openpyxl
    import os
    import platform
    import subprocess
    import json
except ModuleNotFoundError as error:
    messagebox.showerror('Error', error)


class RoomFinanceManager(ttk.Notebook):

    def __init__(self, master, **kw):
        ttk.Notebook.__init__(self, master, **kw)
        self.master = master
        self.heading_font = ('courier', 20, 'bold')
        ttk.Style().configure('TNotebook', background='gray')
        # frames
        self.p1 = ttk.Frame(self, padding=5)
        self.p2 = ttk.Frame(self, padding=5)
        self.p3 = ttk.Frame(self, padding=5)
        self.p4 = ttk.Frame(self, padding=5)
        self.p5 = ttk.Frame(self, padding=5)
        self.add(self.p1, text='Registration')
        self.add(self.p2, text='Student Information')
        self.add(self.p3, text='Treasuring & Auditing')
        self.add(self.p4, text='All Students')
        self.add(self.p5, text='Administration')
        # registration
        self.fullname_frame = ttk.LabelFrame(self.p1, text='Fullname',
                                             padding=5)
        self.first_name = StringVar()
        self.surname = StringVar()
        self.middle_name = StringVar()
        self.age_gender_frame = ttk.LabelFrame(self.p1, text='Age and Gender',
                                               padding=5)
        self.age = StringVar()
        self.gender = StringVar()
        # student information
        self.student_name_to_find = StringVar()
        self.find_entry = ttk.Entry(
            self.p2, textvariable=self.student_name_to_find)
        self.info_frame = ttk.LabelFrame(self.p2, text='Information', padding=5)
        self.pay_frame = ttk.LabelFrame(self.p2, text='Payment', padding=5)
        self.withdraw_frame = ttk.LabelFrame(self.p2, text='Withdrawal',
                                             padding=5)
        self.info_surname = StringVar()
        self.info_fn = StringVar()
        self.info_mn = StringVar()
        self.info_age = StringVar()
        self.info_gender = StringVar()
        self.info_contributed = StringVar()
        self.pay_amount = StringVar()
        self.withdraw_amount = StringVar()
        self.current_student = {}
        # treasuring and auditing
        self.expenses_frame = ttk.LabelFrame(self.p3, text='Expenses',
                                             padding=5)
        self.expense = StringVar()
        self.expense_cost = StringVar()
        self.expense_remove = StringVar()
        self.expenses_list_frame = ttk.LabelFrame(self.p3, text='Expenses List',
                                                  padding=5)
        self.expenses_list = StringVar()
        self.expenses_listbox = Listbox(
            self.expenses_list_frame, listvariable=self.expenses_list, width=40)
        self.expense_s1 = ttk.Scrollbar(
            self.expenses_list_frame, orient=VERTICAL,
            command=self.expenses_listbox.yview)
        self.expense_s2 = ttk.Scrollbar(
            self.expenses_list_frame, orient=HORIZONTAL,
            command=self.expenses_listbox.xview)
        self.expenses_listbox['xscrollcommand'] = self.expense_s2.set
        self.expenses_listbox['yscrollcommand'] = self.expense_s1.set
        self.store_frame = ttk.LabelFrame(self.p3, text='Store', padding=5)
        self.selling_item = StringVar()
        self.cost_sold = StringVar()
        self.item_remove = StringVar()
        self.store_list_frame = ttk.LabelFrame(self.p3, text='Selling List',
                                               padding=5)
        self.store_list = StringVar()
        self.store_listbox = Listbox(
            self.store_list_frame, listvariable=self.store_list, width=40)
        self.selling_s1 = ttk.Scrollbar(
            self.store_list_frame, orient=VERTICAL,
            command=self.store_listbox.yview)
        self.selling_s2 = ttk.Scrollbar(
            self.store_list_frame, orient=HORIZONTAL,
            command=self.store_listbox.xview)
        self.store_listbox['xscrollcommand'] = self.selling_s2.set
        self.store_listbox['yscrollcommand'] = self.selling_s1.set
        self.summary_frame = ttk.LabelFrame(self.p3, text='Summary',
                                            padding=5)
        self.students_fund = StringVar()
        self.total_expenses_cost = StringVar()
        self.total_earned_cost = StringVar()
        self.current_fund = StringVar()
        # all students
        self.boys = []
        self.girls = []
        self.boys_list_frame = ttk.LabelFrame(self.p4, text='Boys List',
                                              padding=5)
        self.boys_list = StringVar()
        self.boys_listbox = Listbox(
            self.boys_list_frame, listvariable=self.boys_list, width=40)
        self.boys_s1 = ttk.Scrollbar(
            self.boys_list_frame, orient=VERTICAL,
            command=self.boys_listbox.yview)
        self.boys_s2 = ttk.Scrollbar(
            self.boys_list_frame, orient=HORIZONTAL,
            command=self.boys_listbox.xview)
        self.boys_listbox['xscrollcommand'] = self.boys_s2.set
        self.boys_listbox['yscrollcommand'] = self.boys_s1.set
        self.girls_list_frame = ttk.LabelFrame(self.p4, text='Girls List',
                                               padding=5)
        self.girls_list = StringVar()
        self.girls_listbox = Listbox(
            self.girls_list_frame, listvariable=self.girls_list, width=40)
        self.girls_s1 = ttk.Scrollbar(
            self.girls_list_frame, orient=VERTICAL,
            command=self.girls_listbox.yview)
        self.girls_s2 = ttk.Scrollbar(
            self.girls_list_frame, orient=HORIZONTAL,
            command=self.girls_listbox.xview)
        self.girls_listbox['xscrollcommand'] = self.girls_s2.set
        self.girls_listbox['yscrollcommand'] = self.girls_s1.set
        self.total_student_frame = ttk.LabelFrame(self.p4, text='Total',
                                                  padding=5)
        self.total_boys = StringVar()
        self.total_girls = StringVar()
        self.total_students = StringVar()
        # administration
        self.admins_frame = ttk.LabelFrame(self.p5, text='Administrators',
                                           padding=5)
        self.admins_list = StringVar()
        self.admins_listbox = Listbox(
            self.admins_frame, listvariable=self.admins_list, width=40)
        self.admins_s1 = ttk.Scrollbar(
            self.admins_frame, orient=VERTICAL,
            command=self.admins_listbox.yview)
        self.admins_s2 = ttk.Scrollbar(
            self.admins_frame, orient=HORIZONTAL,
            command=self.admins_listbox.xview)
        self.admins_listbox['xscrollcommand'] = self.admins_s2.set
        self.admins_listbox['yscrollcommand'] = self.admins_s1.set
        self.login_frame = ttk.LabelFrame(self.p5, text='Login', padding=5)
        self.login_name = StringVar()
        self.login_pass = StringVar()
        self.sign_up_frame = ttk.LabelFrame(self.p5, text='Sign-up', padding=5)
        self.sign_name = StringVar()
        self.sign_pass = StringVar()
        self.remove_frame = ttk.LabelFrame(self.p5, text='Remove', padding=5)
        self.remove_name = StringVar()
        self.other_frame = ttk.LabelFrame(self.p5, text='Others', padding=5)
        # status
        self.status = StringVar()
        self.full_access = False
        self.current_admin = None
        self.generating_report = False
        # geometry management
        for i in range(9):  # row
            self.rowconfigure(i, weight=1)
            self.p1.rowconfigure(i, weight=1)
            self.p2.rowconfigure(i, weight=1)
            self.p3.rowconfigure(i, weight=1)
            self.p4.rowconfigure(i, weight=1)
            self.p5.rowconfigure(i, weight=1)
            if i <= 1:
                self.login_frame.rowconfigure(i, weight=1)
                self.sign_up_frame.rowconfigure(i, weight=1)
                self.remove_frame.rowconfigure(i, weight=1)
                self.other_frame.rowconfigure(i, weight=1)
                self.age_gender_frame.rowconfigure(i, weight=1)
                self.expenses_frame.columnconfigure(i, weight=1)
                if i != 1:
                    self.expenses_list_frame.columnconfigure(i, weight=1)
                    self.store_list_frame.columnconfigure(i, weight=1)
                self.store_frame.columnconfigure(i, weight=1)
            if i <= 6:
                if i != 6:
                    self.admins_frame.rowconfigure(i, weight=1)
                self.info_frame.rowconfigure(i, weight=1)
                if i <= 2:
                    self.fullname_frame.rowconfigure(i, weight=1)
                if i <= 3:
                    self.pay_frame.rowconfigure(i, weight=1)
                    self.withdraw_frame.rowconfigure(i, weight=1)
                    self.expenses_frame.rowconfigure(i, weight=1)
                    if i not in [2, 3]:
                        self.expenses_list_frame.rowconfigure(i, weight=1)
                        self.store_list_frame.rowconfigure(i, weight=1)
                    self.store_frame.rowconfigure(i, weight=1)
                    self.total_student_frame.rowconfigure(i, weight=1)
                    self.summary_frame.rowconfigure(i, weight=1)
                    self.boys_list_frame.rowconfigure(i, weight=1)
                    self.girls_list_frame.rowconfigure(i, weight=1)
        for i in range(4):  # column
            self.columnconfigure(i, weight=1)
            self.p1.columnconfigure(i, weight=1)
            self.p2.columnconfigure(i, weight=1)
            self.p3.columnconfigure(i, weight=1)
            self.p4.columnconfigure(i, weight=1)
            self.p5.columnconfigure(i, weight=1)
            if i <= 1:
                self.login_frame.columnconfigure(i, weight=1)
                self.sign_up_frame.columnconfigure(i, weight=1)
                self.remove_frame.columnconfigure(i, weight=1)
                self.other_frame.columnconfigure(i, weight=1)
                if i != 1:
                    self.admins_frame.columnconfigure(i, weight=1)
                self.info_frame.columnconfigure(i, weight=1)
                self.pay_frame.columnconfigure(i, weight=1)
                self.withdraw_frame.columnconfigure(i, weight=1)
                if i != 1:
                    self.boys_list_frame.columnconfigure(i, weight=1)
                    self.girls_list_frame.columnconfigure(i, weight=1)
            if i <= 4:
                self.summary_frame.columnconfigure(i, weight=1)
                self.total_student_frame.columnconfigure(i, weight=1)
            self.fullname_frame.columnconfigure(i, weight=1)
            self.age_gender_frame.columnconfigure(i, weight=1)
        # file initialization
        if os.path.exists('data.txt'):
            with open('data.txt', 'r') as d:
                self.data = json.load(d)
        else:
            self.data = {
                'admins': {
                    'Admin': encrypt('code')
                },
                'boys': {},
                'girls': {},
                'expenses': {},
                'store': {}
            }
            self.update_data()
            messagebox.showinfo(
                'Info', 'This is your first time using the application. '
                        'The temporary admin is "Admin", password is '
                        '"code". Please register a new admin.')
        self.select(4)
        # finding excel path
        self.excel_path = ''
        if platform.machine() == 'x86':
            self.excel_path = 'C:\\Program Files\\Microsoft Office\\'
        else:
            self.excel_path = 'C:\\Program Files (x86)\\Microsoft Office\\'
        for i in os.listdir(self.excel_path):
            if i.startswith('Office'):
                self.excel_path += i
                break
        self.excel_path += '\\EXCEL.EXE'
        # events
        self.master.bind('<Enter>', self.update_app)
        self.find_entry.bind('<Key>', self.find_name)
        self.find_entry.bind('<Return>', self.find_name)
        self.boys_listbox.bind('<<ListboxSelect>>', self.pass_boy)
        self.girls_listbox.bind('<<ListboxSelect>>', self.pass_girl)
        self.boys_listbox.bind('<Return>', self.find_name)
        self.girls_listbox.bind('<Return>', self.find_name)
        self.boys_listbox.bind('<Double-1>', self.find_name)
        self.girls_listbox.bind('<Double-1>', self.find_name)
        self.master.bind('<F1>', self.show_about_dialog)

        self.create_widgets()

    def show_about_dialog(self, *args):
        d = AboutDialog(
            self, window_title='About Room Finance Manager',
            about_title='Room Finance Manager',
            content='Developed and written by:\n'
                    '\tDenniel Luis Saway Sadian '
                    '(https://denniel-sadian.github.io)\n\n'
                    'Date of creation:\n'
                    '\tFebruary 18, 2018\n\n'
                    'Description:\n'
                    '\tThis application is developed for finance management in '
                    'school rooms only. It can generate reports in excel '
                    'spread sheets. It also makes treasuring and auditing '
                    'simple and erroneous-free.',
            image='rfm.png')
        d.wm_iconbitmap('icon.ico')
        d.mainloop()
        return args

    def modify_info(self):
        if all([self.info_surname.get(), self.info_fn.get(), self.info_mn.get(),
                self.info_age.get(), self.info_gender.get()]):
            sn = self.info_surname.get().strip()
            fn = self.info_fn.get().strip()
            mn = self.info_mn.get().strip()
            gender = self.info_gender.get().lower().strip()
            age = int(self.info_age.get().strip())
            name = f'{sn}, {fn} {mn}'
            if gender in ['boy', 'girl']:
                if self.full_access:
                    del self.data[self.current_student['gender'] + 's'][
                        self.current_student['key']]
                    student = {
                        'fn': fn,
                        'mn': mn,
                        'sn': sn,
                        'age': age,
                        'gender': gender,
                        'paid': self.current_student['paid']
                    }
                    self.data[gender + 's'][name] = student
                    self.update_data()
                    messagebox.showinfo(
                        'Info', f"The data of {student['fn']} {student['sn']} "
                                "have been modified, expect the contribution.")
                else:
                    messagebox.showinfo(
                        'Info', 'You cannot modify these data unless an admin '
                                'is logged in.')
            else:
                messagebox.showerror(
                    'Error', 'Boy and girl are the only accepted gender.')
                self.info_gender.set('Boy or Girl?')
            self.student_name_to_find.set(name)
            self.find_name()

    def truly_generate_report(self):
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        # heading
        sheet.merge_cells('A1:B1')
        sheet.merge_cells('C1:E1')
        sheet.merge_cells('C2:E2')
        sheet.merge_cells('A2:B2')
        sheet['A1'] = 'Generated by:'
        sheet['C1'] = self.current_admin
        sheet['A2'] = 'Date and Time:'
        sheet['C2'] = strftime('%m/%d/%Y - %I:%M %p')
        for cell in ['A1', 'A2', 'A4', 'B4', 'D4', 'E4', 'G4', 'H4', 'I4',
                     'K4', 'L4', 'M4', 'O4', 'O5', 'O6', 'O8', 'O9', 'O10',
                     'O11']:
            if cell in 'A1 A2':
                sheet[cell].alignment = Alignment(horizontal='right')
            else:
                sheet[cell].alignment = Alignment(horizontal='center')
            sheet[cell].fill = PatternFill('solid', Color('ffff00'))
        for cell in ['C1', 'C2']:
            sheet[cell].alignment = Alignment(horizontal='center')
            sheet[cell].fill = PatternFill('solid', Color('ffff80'))
        # boys and girls
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet['A4'] = 'Boys'
        sheet['B4'] = 'Paid'
        sheet['D4'] = 'Girls'
        sheet['E4'] = 'Paid'
        boys = sorted(list(self.data['boys']))
        for i in range(len(boys)):
            cell1 = 'A' + str(i + 5)
            cell2 = 'B' + str(i + 5)
            sheet[cell1] = boys[i]
            sheet[cell2] = self.data['boys'][boys[i]]['paid']
            sheet[cell1].alignment = Alignment(horizontal='left')
            sheet[cell2].alignment = Alignment(horizontal='center')
        girls = sorted(list(self.data['girls']))
        for i in range(len(girls)):
            cell1 = 'D' + str(i + 5)
            cell2 = 'E' + str(i + 5)
            sheet[cell1] = girls[i]
            sheet[cell2] = self.data['girls'][girls[i]]['paid']
            sheet[cell1].alignment = Alignment(horizontal='left')
            sheet[cell2].alignment = Alignment(horizontal='center')
        # expenses and selling
        for col in ['G', 'I', 'K', 'M']:
            sheet.column_dimensions[col].width = 15
        for i in [('G4', 'Expenses'), ('H4', 'Cost'), ('I4', 'Date'),
                  ('K4', 'Selling Items'), ('L4', 'Earned'),
                  ('M4', 'Date'), ('O4', 'Boys:'), ('O5', 'Girls:'),
                  ('O6', 'All:'), ('O8', 'Class Fund:'),
                  ('O9', 'Total Expenses:'), ('O10', 'Total Earned:'),
                  ('O11', 'Current Class Funds:')]:
            sheet[i[0]] = i[1]
        expenses = list(self.data['expenses'])
        for i in range(len(expenses)):
            cell1 = 'G' + str(i + 5)
            cell2 = 'H' + str(i + 5)
            cell3 = 'I' + str(i + 5)
            sheet[cell1] = expenses[i]
            sheet[cell2] = self.data['expenses'][expenses[i]]['cost']
            sheet[cell3] = self.data['expenses'][expenses[i]]['date']
            sheet[cell1].alignment = Alignment(horizontal='left')
            sheet[cell2].alignment = Alignment(horizontal='center')
            sheet[cell3].alignment = Alignment(horizontal='center')
        selling = list(self.data['store'])
        for i in range(len(selling)):
            cell1 = 'K' + str(i + 5)
            cell2 = 'L' + str(i + 5)
            cell3 = 'M' + str(i + 5)
            sheet[cell1] = selling[i]
            sheet[cell2] = self.data['store'][selling[i]]['earned']
            sheet[cell3] = self.data['store'][selling[i]]['date']
            sheet[cell1].alignment = Alignment(horizontal='left')
            sheet[cell2].alignment = Alignment(horizontal='center')
            sheet[cell3].alignment = Alignment(horizontal='center')
        # summary
        sheet.column_dimensions['O'].width = 20
        for i in [('P4', int(self.total_boys.get())),
                  ('P5', int(self.total_girls.get())),
                  ('P6', int(self.total_students.get())),
                  ('P8', float(self.students_fund.get())),
                  ('P9', float(self.total_expenses_cost.get())),
                  ('P10', float(self.total_earned_cost.get())),
                  ('P11', float(self.current_fund.get()))]:
            sheet[i[0]] = i[1]
            sheet[i[0]].alignment = Alignment(horizontal='center')
            sheet[i[0]].fill = PatternFill('solid', Color('ffff80'))
        file_name = strftime('%B %d, %Y.xlsx')
        wb.save(file_name)
        subprocess.Popen([self.excel_path, os.path.abspath(file_name)])
        self.generating_report = False
        messagebox.showinfo('Info', 'Done!')

    def generate_report(self):
        if self.full_access:
            if not self.generating_report:
                self.generating_report = True
                Thread(target=self.truly_generate_report, args=()).start()
            else:
                messagebox.showinfo(
                    'Info', 'The application is already generating a report.')
        else:
            messagebox.showinfo(
                'Info', "You are not authorized to generate a report unless "
                        "you are the admin and you logged in to this app. The "
                        "admin's name is needed in generating a report.")

    def update_data(self):
        with open('data.txt', 'w') as d:
            json.dump(self.data, d, indent=4)

    def login(self):
        if all([self.login_name.get(), self.login_pass.get()]):
            name = self.login_name.get()
            _pass = self.login_pass.get()
            if name in self.data['admins']:
                if decrypt(self.data['admins'][name]) == _pass:
                    self.full_access = True
                    self.current_admin = name
                    messagebox.showinfo(
                        'Welcome', f'Welcome administrator {name}!')
                else:
                    messagebox.showerror('Error', 'Wrong password!')
            else:
                messagebox.showinfo(
                    'Info', f'"{name}" is not an administrator.')
            self.login_name.set('')
            self.login_pass.set('')

    def sign_up(self):
        if all([self.sign_name.get(), self.sign_pass.get()]):
            if self.full_access and self.current_admin:
                name = self.sign_name.get()
                _pass = self.sign_pass.get()
                if self.sign_name.get() not in self.data['admins']:
                    self.data['admins'][name] = encrypt(_pass)
                    self.update_data()
                    messagebox.showinfo(
                        'Info', f'{name} has been registered as an admin by '
                                f'administrator {self.current_admin}')
                else:
                    messagebox.showinfo(
                        'Info', f'{name} has been registered already.')
            else:
                messagebox.showinfo('Info', 'Access denied. There must be an '
                                            'administrator logged in.')
            self.sign_name.set('')
            self.sign_pass.set('')

    def remove_admin(self):
        if self.remove_name.get():
            name = self.remove_name.get()
            if self.full_access and self.current_admin:
                if len(self.data['admins']) == 1:
                    messagebox.showinfo(
                        'Info', 'There is only one admin registered in the '
                                'app, you cannot remove an admin at this '
                                'moment.')
                elif name in self.data['admins']:
                    del self.data['admins'][name]
                    self.update_data()
                    messagebox.showinfo(
                        'Info', f'"{name}" has been removed as an admin by '
                                f'admin {self.current_admin}')
                    if name == self.current_admin:
                        self.lock_access()
                else:
                    messagebox.showinfo('Info', f'"{name}" is not an admin.')
            else:
                messagebox.showinfo('Info',
                                    'Access denied. You cannot remove an admin '
                                    'unless there is an admin logged in.')
            self.remove_name.set('')

    def lock_access(self):
        if self.full_access:
            self.full_access = False
            self.current_admin = None
            self.update_app()
            messagebox.showinfo('Info', 'Access has been limited.')
        else:
            messagebox.showinfo('Info', 'Access is limited already.')

    def wipe_data(self):
        if self.full_access:
            if messagebox.askokcancel(
                    'Wiping data', 'Are you sure that you want to wipe the '
                                   'data? After doing so, the current data of '
                                   'this application will be lost forever. '
                                   'Proceed?'):
                self.data = {
                    'admins': {
                        self.current_admin: self.data['admins']
                        [self.current_admin]
                    },
                    'boys': {},
                    'girls': {},
                    'expenses': {},
                    'store': {}
                }
                self.update_data()
                messagebox.showwarning(
                    'Info', 'Data has been wiped completely.')
        else:
            messagebox.showinfo('Info', 'Access denied!')

    def pay(self):
        if all([self.info_surname.get(), self.info_fn.get(), self.info_mn.get(),
                self.pay_amount.get(), self.info_gender.get()]):
            name = f"{self.info_surname.get()}, {self.info_fn.get()} " \
                   f"{self.info_mn.get()}"
            try:
                if float(self.pay_amount.get()) > 0:
                    self.data[self.info_gender.get() + 's'][name]['paid'] \
                        += float(self.pay_amount.get())
                    self.update_data()
                    self.find_name()
                    self.pay_amount.set('')
                    messagebox.showinfo(
                        'Info', f'Thanks for paying {self.info_fn.get()} '
                                f'{self.info_surname.get()}!')
                else:
                    messagebox.showinfo('Info', 'Cannot be taken')
                    self.pay_amount.set('')
            except ValueError:
                messagebox.showerror('Error', 'Value Error')
                self.pay_amount.set('')
        else:
            self.pay_amount.set('')

    def withdraw(self):
        if all([self.info_surname.get(), self.info_fn.get(), self.info_mn.get(),
                self.withdraw_amount.get(), self.info_gender.get()]):
            if self.full_access:
                name = f"{self.info_surname.get()}, {self.info_fn.get()} " \
                       f"{self.info_mn.get()}"
                try:
                    paid = self.data[self.info_gender.get() + 's'][name]['paid']
                    if 0 < float(self.withdraw_amount.get()) <= float(paid):
                        self.data[self.info_gender.get() + 's'][name]['paid'] \
                            -= float(self.withdraw_amount.get())
                        self.update_data()
                        self.find_name()
                        self.withdraw_amount.set('')
                        messagebox.showinfo(
                            'Info', f'Hope you can pay for that '
                                    f'{self.info_fn.get()} '
                                    f'{self.info_surname.get()}.')
                    else:
                        messagebox.showinfo('Info', 'Cannot be taken')
                        self.withdraw_amount.set('')
                except ValueError:
                    messagebox.showerror('Error', 'Value Error')
                    self.withdraw_amount.set('')
            else:
                messagebox.showinfo(
                    'Info', 'You cannot withdraw unless there is an '
                            'admin logged in.')
        else:
            self.pay_amount.set('')

    def add_or_modify(self):
        if all([self.expense.get(), self.expense_cost.get()]):
            if self.full_access:
                try:
                    if 0 < float(self.expense_cost.get()) < \
                            float(self.current_fund.get()):
                        self.data['expenses'][self.expense.get().lower()] = {
                            'cost': float(self.expense_cost.get()),
                            'date': strftime('%m/%d/%Y')
                        }
                        self.update_data()
                        self.update_app()
                        messagebox.showinfo(
                            'Info', f'"{self.expense.get()}" has been '
                                    f'added/modified.')
                    else:
                        messagebox.showinfo(
                            'Info', "Cannot add or modify an item at this "
                                    "moment. Please check the class' current "
                                    "fund or amount of the entry.")
                except ValueError:
                    messagebox.showerror('Error', 'Value Error!')
            else:
                messagebox.showinfo(
                    'Info', 'You cannot add or modify an item unless there is '
                            'an admin logged in.')
            self.expense.set('')
            self.expense_cost.set('')

    def remove_expense(self):
        if self.expense_remove.get():
            if self.full_access:
                item = self.expense_remove.get().lower()
                if item in self.data['expenses']:
                    if messagebox.askyesno(
                            'Info', 'Are you sure you want to remove '
                                    f'"{item}" from the list?'):
                        del self.data['expenses'][item]
                        self.update_data()
                else:
                    messagebox.showinfo('Info', f'"{item}" is not in the list.')
            else:
                messagebox.showinfo(
                    'Info', 'You cannot remove an item unless there is '
                            'an admin logged in.')
            self.expense_remove.set('')

    def add_or_modify_selling(self):
        if all([self.selling_item.get(), self.cost_sold.get()]):
            if self.full_access:
                try:
                    if float(self.cost_sold.get()) > 0:
                        item = self.selling_item.get().lower()
                        earned = float(self.cost_sold.get())
                        if item in self.data['store']:
                            self.data['store'][item]['earned'] += earned
                            self.update_data()
                            messagebox.showinfo(
                                'Info', f'"{item}" has been modified.')
                        else:
                            self.data['store'][item] = {
                                'earned': earned,
                                'date': strftime('%m/%d/%Y')
                            }
                            self.update_data()
                            messagebox.showinfo('Info', f'"{item}" has been '
                                                        f'added from the list.')
                except ValueError:
                    messagebox.showerror('Error', 'Value Error!')
            else:
                messagebox.showinfo(
                    'Info', 'You cannot add or update an item unless there is '
                            'an admin logged in.')

    def remove_selling(self):
        if self.item_remove.get():
            if self.full_access:
                item = self.item_remove.get().lower()
                if item in self.data['store']:
                    if messagebox.askyesno(
                            'Info', 'Are you sure you want to remove '
                                    f'"{item}" from the list?'):
                        del self.data['store'][item]
                        self.update_data()
                else:
                    messagebox.showinfo('Info', f'"{item}" is not in the list.')
            else:
                messagebox.showinfo(
                    'Info', 'You cannot remove an item unless there is '
                            'an admin logged in.')
            self.item_remove.set('')

    def search(self, name):
        for gender in ['boys', 'girls']:
            for i in self.data[gender].keys():
                if name.lower() in i.lower():
                    student = self.data[gender][i]
                    student['key'] = i
                    return student
        return None

    def find_name(self, *args):
        if self.student_name_to_find.get():
            if self.search(self.student_name_to_find.get()):
                self.select(1)
                self.current_student = self.search(
                    self.student_name_to_find.get())
                self.info_surname.set(self.current_student['sn'])
                self.info_fn.set(self.current_student["fn"])
                self.info_mn.set(self.current_student['mn'])
                self.info_age.set(self.current_student['age'])
                self.info_gender.set(self.current_student['gender'])
                self.info_contributed.set(
                    self.current_student['paid'])
        else:
            for i in [self.info_surname, self.info_fn,
                      self.info_mn, self.info_age, self.info_gender,
                      self.info_contributed]:
                i.set('')
        return args

    def pass_boy(self, *args):
        if len(self.boys_listbox.curselection()) == 1:
            self.student_name_to_find.set(sorted(self.boys)[int(
                self.boys_listbox.curselection()[0])])
        return args

    def pass_girl(self, *args):
        if len(self.girls_listbox.curselection()) == 1:
            self.student_name_to_find.set(sorted(self.girls)[int(
                self.girls_listbox.curselection()[0])])
        return args

    def update_app(self, *args):
        # setting strings
        self.total_boys.set(len(list(self.data['boys'])))
        self.total_girls.set(len(list(self.data['girls'])))
        self.total_students.set(int(self.total_boys.get()) +
                                int(self.total_girls.get()))
        fund = 0
        if int(self.total_boys.get()) > 0:
            for i in self.data['boys']:
                fund += self.data['boys'][i]['paid']
        if int(self.total_girls.get()) > 0:
            for i in self.data['girls']:
                fund += self.data['girls'][i]['paid']
        self.students_fund.set(fund)
        total_expenses = 0
        if len(self.data['expenses']) > 0:
            for i in self.data['expenses']:
                total_expenses += self.data['expenses'][i]['cost']
        self.total_expenses_cost.set(total_expenses)
        total_earned = 0
        if len(self.data['store']) > 0:
            for i in self.data['store']:
                total_earned += self.data['store'][i]['earned']
        self.total_earned_cost.set(total_earned)
        self.current_fund.set((fund-total_expenses) + total_earned)
        self.status.set(f'Full access: {self.full_access} | ')
        self.status.set(self.status.get() + f'Admin: {self.current_admin}')
        # setting lists to string variables
        self.admins_list.set(list(self.data['admins']))
        self.boys = []
        for i in self.data['boys'].keys():
            name = f"{self.data['boys'][i]['sn']}, " \
                   f"{self.data['boys'][i]['fn']} {self.data['boys'][i]['mn']}"
            self.boys.append(name)
        self.boys_list.set(sorted(self.boys))
        self.girls = []
        for i in self.data['girls'].keys():
            name = f"{self.data['girls'][i]['sn']}, " \
                   f"{self.data['girls'][i]['fn']} {self.data['girls'][i]['mn']}"
            self.girls.append(name)
        self.girls_list.set(sorted(self.girls))
        expenses = []
        for i in self.data['expenses']:
            name = f"{self.data['expenses'][i]['date']}  {i.upper()}  " \
                   f"{self.data['expenses'][i]['cost']}"
            expenses.append(name)
        selling = []
        for i in self.data['store']:
            name = f"{self.data['store'][i]['date']}  {i.upper()}  " \
                   f"{self.data['store'][i]['earned']}"
            selling.append(name)
        self.expenses_list.set(expenses)
        self.store_list.set(selling)
        # styling list boxes
        for listbox, items in \
                [(self.admins_listbox, len(list(self.data['admins']))),
                 (self.boys_listbox, len(list(self.data['boys']))),
                 (self.girls_listbox, len(list(self.data['girls']))),
                 (self.expenses_listbox, len(list(self.data['expenses']))),
                 (self.store_listbox, len(list(self.data['store'])))]:
            for i in range(0, items, 2):
                listbox.itemconfigure(i, background='#f0f0ff')
        return args

    def register_student(self):
        if all([self.first_name.get(), self.surname.get(),
                self.middle_name.get(), self.age.get(),
                self.gender.get()]):
            if self.full_access:
                try:
                    surname = self.surname.get().strip()
                    first_name = self.first_name.get().strip()
                    middle_name = self.middle_name.get().strip()
                    age = int(self.age.get().strip())
                    gender = self.gender.get()
                    key = f'{surname}, {first_name} {middle_name}'
                    student = {
                        'fn': first_name,
                        'mn': middle_name,
                        'sn': surname,
                        'age': age,
                        'gender': gender,
                        'paid': 0
                    }
                    if not self.search(key):
                        self.data[student['gender'] + 's'][key] = student
                        self.update_data()
                        messagebox.showinfo(
                            'Info', f'{key} has been registered.')
                    else:
                        if messagebox.askyesno(
                                'Info', 'This student has been registered '
                                        'already. Do you want to modify the '
                                        'data?'):
                            self.student_name_to_find.set(key)
                            self.find_name()
                            self.select(1)
                    for i in [self.first_name, self.surname,
                              self.middle_name, self.age, self.gender]:
                        i.set('')
                except ValueError:
                    messagebox.showerror('Error', 'Invalid age value.')
            else:
                messagebox.showinfo(
                    'Info', 'You cannot register a student unless an '
                            'admin is logged in.')
        else:
            messagebox.showinfo('Error', 'Incomplete.')

    def create_widgets(self):
        # registration
        ttk.Label(self.p1, text='Registration',
                  font=self.heading_font).grid(
            column=0, row=0, columnspan=4, pady='10 0')
        ttk.Separator(self.p1, orient=HORIZONTAL).grid(
            column=0, row=1, sticky='WE', columnspan=4, pady=5)
        # ---- fullname
        self.fullname_frame.grid(column=0, row=2, sticky='NEWS', pady=5, padx=5,
                                 columnspan=4, rowspan=3)
        ttk.Label(self.fullname_frame, text='First name').grid(column=0, row=0)
        ttk.Entry(self.fullname_frame, textvariable=self.first_name).grid(
            column=1, row=0, sticky='WE', columnspan=3)
        ttk.Label(self.fullname_frame, text='Surname').grid(column=0, row=1)
        ttk.Entry(self.fullname_frame, textvariable=self.surname).grid(
            column=1, row=1, sticky='WE', columnspan=3)
        ttk.Label(self.fullname_frame, text='Middle name').grid(column=0, row=2)
        ttk.Entry(self.fullname_frame, textvariable=self.middle_name).grid(
            column=1, row=2, sticky='WE', columnspan=3)
        # ---- age and gender
        self.age_gender_frame.grid(column=0, row=5, sticky='NEWS', pady=5,
                                   padx=5, columnspan=4, rowspan=2)
        ttk.Label(self.age_gender_frame, text='Age').grid(column=0, row=0)
        ttk.Entry(self.age_gender_frame, textvariable=self.age).grid(
            column=1, row=0, sticky='WE', columnspan=3)
        ttk.Checkbutton(self.age_gender_frame, text='Boy', variable=self.gender,
                        onvalue='boy', offvalue='girl').grid(
            column=1, row=1, sticky=W)
        ttk.Checkbutton(self.age_gender_frame, text='Girl',
                        variable=self.gender, onvalue='girl',
                        offvalue='boy').grid(column=2, row=1, sticky=W)
        ttk.Separator(self.p1, orient=HORIZONTAL).grid(
            column=0, row=7, sticky='WE', columnspan=4, pady=3)
        ttk.Button(self.p1, text='Register', command=self.register_student).grid(
            column=0, row=8, sticky='NEWS', columnspan=4, rowspan=2)
        # student information
        ttk.Label(self.p2, text='Student Information',
                  font=self.heading_font).grid(
            column=0, row=0, columnspan=4, pady='10 0')
        ttk.Label(self.p2, text='Name').grid(column=0, row=2)
        self.find_entry.grid(column=1, row=2, sticky='WE', columnspan=3)
        ttk.Separator(self.p2, orient=HORIZONTAL).grid(
            column=0, row=3, sticky='WE', columnspan=4)
        ttk.Separator(self.p2, orient=HORIZONTAL).grid(
            column=0, row=1, sticky='WE', columnspan=4, pady=5)
        self.info_frame.grid(column=0, row=4, sticky='NEWS', columnspan=2,
                             rowspan=6, pady=5, padx=5)
        ttk.Label(self.info_frame, text='Surname').grid(
            column=0, row=0)
        ttk.Entry(self.info_frame, textvariable=self.info_surname).grid(
            column=1, row=0, sticky='WE')
        ttk.Label(self.info_frame, text='First name').grid(
            column=0, row=1)
        ttk.Entry(self.info_frame, textvariable=self.info_fn).grid(
            column=1, row=1, sticky='WE')
        ttk.Label(self.info_frame, text='Middle name').grid(
            column=0, row=2)
        ttk.Entry(self.info_frame, textvariable=self.info_mn).grid(
            column=1, row=2, sticky='WE')
        ttk.Label(self.info_frame, text='Age').grid(
            column=0, row=3)
        ttk.Entry(self.info_frame, textvariable=self.info_age).grid(
            column=1, row=3, sticky='WE')
        ttk.Label(self.info_frame, text='Gender').grid(
            column=0, row=4)
        ttk.Entry(self.info_frame, textvariable=self.info_gender).grid(
            column=1, row=4, sticky='WE')
        ttk.Label(self.info_frame, text='Contributed').grid(
            column=0, row=5)
        ttk.Entry(self.info_frame, textvariable=self.info_contributed).grid(
            column=1, row=5, sticky='WE')
        ttk.Button(self.info_frame, text='Modify',
                   command=self.modify_info).grid(
            column=0, row=6, sticky='NEWS', columnspan=2)
        # ---- pay frame
        self.pay_frame.grid(column=2, row=4, sticky='NEWS', columnspan=2,
                            rowspan=3, pady=5, padx=5)
        ttk.Label(self.pay_frame, text='Please pay').grid(
            column=0, row=0, columnspan=2, pady='0 5')
        ttk.Label(self.pay_frame, text='Amount').grid(column=0, row=1)
        ttk.Entry(self.pay_frame, textvariable=self.pay_amount).grid(
            column=1, row=1, sticky='WE')
        ttk.Button(self.pay_frame, text='Pay', command=self.pay).grid(
            column=1, row=2, sticky='NEWS')
        # ---- withdraw frame
        self.withdraw_frame.grid(column=2, row=7, sticky='NEWS', columnspan=2,
                                 rowspan=3, pady=5, padx=5)
        ttk.Label(self.withdraw_frame, text='You can also withdraw').grid(
            column=0, row=0, columnspan=2, pady='0 5')
        ttk.Label(self.withdraw_frame, text='Amount').grid(column=0, row=1)
        ttk.Entry(self.withdraw_frame, textvariable=self.withdraw_amount).grid(
            column=1, row=1, sticky='WE')
        ttk.Button(self.withdraw_frame, text='Withdraw',
                   command=self.withdraw).grid(column=1, row=2, sticky='NEWS')
        # treasuring and auditing
        ttk.Label(self.p3, text='Treasuring and Auditing',
                  font=self.heading_font).grid(
            column=0, row=0, columnspan=4, pady='10 0')
        ttk.Separator(self.p3, orient=HORIZONTAL).grid(
            column=0, row=1, sticky='WE', columnspan=4, pady=5)
        # ---- expenses
        self.expenses_frame.grid(column=0, row=2, sticky='NEWS', columnspan=2,
                                 rowspan=3, pady=5, padx=5)
        ttk.Label(self.expenses_frame, text='Bought').grid(column=0, row=0)
        ttk.Entry(self.expenses_frame, textvariable=self.expense).grid(
            column=1, row=0, sticky='WE')
        ttk.Label(self.expenses_frame, text='Cost').grid(column=0, row=1)
        ttk.Entry(self.expenses_frame, textvariable=self.expense_cost).grid(
            column=1, row=1, sticky='WE')
        ttk.Button(self.expenses_frame, text='Add / Modify',
                   command=self.add_or_modify).grid(
                       column=1, row=2, sticky='WE')
        ttk.Label(self.expenses_frame, text='Remove').grid(column=0, row=3)
        ttk.Entry(self.expenses_frame, textvariable=self.expense_remove).grid(
            column=1, row=3, sticky='WE', pady='8 0')
        ttk.Button(self.expenses_frame, text='Remove',
                   command=self.remove_expense).grid(
            column=1, row=4, sticky='WE')
        # ---- expenses list
        self.expenses_list_frame.grid(column=0, row=5, sticky='NEWS',
                                      columnspan=2, rowspan=3, pady=5, padx=5)
        self.expenses_listbox.grid(column=0, row=0, sticky='NEWS', rowspan=2)
        self.expense_s1.grid(column=1, row=0, sticky='NSW', rowspan=2)
        self.expense_s2.grid(column=0, row=2, sticky='WEN')
        # ---- store
        self.store_frame.grid(column=2, row=2, sticky='NEWS', columnspan=2,
                              rowspan=3, pady=5, padx=5)
        ttk.Label(self.store_frame, text='Selling item').grid(column=0, row=0)
        ttk.Entry(self.store_frame, textvariable=self.selling_item).grid(
            column=1, row=0, sticky='WE')
        ttk.Label(self.store_frame, text='Cost sold').grid(column=0, row=1)
        ttk.Entry(self.store_frame, textvariable=self.cost_sold).grid(
            column=1, row=1, sticky='WE')
        ttk.Button(self.store_frame, text='Add / Update',
                   command=self.add_or_modify_selling).grid(
            column=1, row=2, sticky='WE')
        ttk.Label(self.store_frame, text='Remove').grid(column=0, row=3)
        ttk.Entry(self.store_frame, textvariable=self.item_remove).grid(
            column=1, row=3, sticky='WE', pady='8 0')
        ttk.Button(self.store_frame, text='Remove',
                   command=self.remove_selling).grid(
            column=1, row=4, sticky='WE')
        # ---- selling list
        self.store_list_frame.grid(column=2, row=5, sticky='NEWS',
                                   columnspan=2, rowspan=3, pady=5, padx=5)
        self.store_listbox.grid(column=0, row=0, sticky='NEWS', rowspan=2)
        self.selling_s1.grid(column=1, row=0, sticky='NSW', rowspan=2)
        self.selling_s2.grid(column=0, row=2, sticky='WEN')
        # ---- summary
        self.summary_frame.grid(column=0, row=8, sticky='NEWS', rowspan=2,
                                columnspan=4, pady=5, padx=5)
        ttk.Label(self.summary_frame, text="Student's fund").grid(
            column=0, row=0)
        ttk.Entry(self.summary_frame, textvariable=self.students_fund).grid(
            column=1, row=0, sticky='WE', columnspan=2)
        ttk.Label(self.summary_frame, text='<--- Got from students').grid(
            column=3, row=0)
        ttk.Label(self.summary_frame, text='Total expenses cost').grid(
            column=0, row=1)
        ttk.Entry(self.summary_frame,
                  textvariable=self.total_expenses_cost).grid(
            column=1, row=1, sticky='WE', columnspan=2)
        ttk.Label(self.summary_frame, text='<--- Got from expenses cost').grid(
            column=3, row=1)
        ttk.Label(self.summary_frame, text='Total earned amount').grid(
            column=0, row=2)
        ttk.Entry(self.summary_frame, textvariable=self.total_earned_cost).grid(
            column=1, row=2, sticky='WE', columnspan=2)
        ttk.Label(self.summary_frame, text='<--- Got from selling items').grid(
            column=3, row=2)
        ttk.Label(self.summary_frame, text='Current class fund').grid(
            column=0, row=3)
        ttk.Entry(self.summary_frame, textvariable=self.current_fund).grid(
            column=1, row=3, sticky='WE', columnspan=2)
        ttk.Label(self.summary_frame, text='<--- The total').grid(
            column=3, row=3)
        # all students
        ttk.Label(self.p4, text='All Students',
                  font=self.heading_font).grid(
            column=0, row=0, columnspan=4, pady='10 0')
        ttk.Separator(self.p4, orient=HORIZONTAL).grid(
            column=0, row=1, sticky='WE', columnspan=4, pady=5)
        # ---- boys list
        self.boys_list_frame.grid(column=0, row=2, sticky='NEWS', columnspan=2,
                                  rowspan=5, pady=5, padx=5)
        self.boys_listbox.grid(column=0, row=0, sticky='NEWS', rowspan=4)
        self.boys_s1.grid(column=1, row=0, sticky='NSW', rowspan=4)
        self.boys_s2.grid(column=0, row=4, sticky='WEN')
        # ---- girls list
        self.girls_list_frame.grid(column=2, row=2, sticky='NEWS', columnspan=2,
                                   rowspan=5, pady=5, padx=5)
        self.girls_listbox.grid(column=0, row=0, sticky='NEWS', rowspan=4)
        self.girls_s1.grid(column=1, row=0, sticky='NSW', rowspan=4)
        self.girls_s2.grid(column=0, row=4, sticky='WEN')
        # ---- total students
        self.total_student_frame.grid(column=0, row=7, sticky='NEWS', pady=5,
                                      padx=5, rowspan=3, columnspan=4)
        ttk.Label(self.total_student_frame, text='Boys').grid(
            column=0, row=0, columnspan=2)
        ttk.Entry(self.total_student_frame, textvariable=self.total_boys).grid(
            column=2, row=0, sticky='WE')
        ttk.Label(self.total_student_frame, text='Girls').grid(
            column=0, row=1, columnspan=2)
        ttk.Entry(self.total_student_frame, textvariable=self.total_girls).grid(
            column=2, row=1, sticky='WE')
        ttk.Label(self.total_student_frame, text='Total students').grid(
            column=0, row=2, columnspan=2)
        ttk.Entry(self.total_student_frame,
                  textvariable=self.total_students).grid(column=2, row=2,
                                                         sticky='WE')
        # administration
        ttk.Label(self.p5, text='Administration',
                  font=self.heading_font).grid(
            column=0, row=0, columnspan=4, pady='10 0')
        ttk.Separator(self.p5, orient=HORIZONTAL).grid(
            column=0, row=1, sticky='WE', columnspan=4, pady=5)
        # ---- login
        self.login_frame.grid(column=2, row=2, sticky='NEWS', columnspan=2,
                              rowspan=3, pady=5, padx=5)
        ttk.Label(self.login_frame, text='Name').grid(column=0, row=0)
        ttk.Entry(self.login_frame, textvariable=self.login_name).grid(
            column=1, row=0, sticky='WE')
        ttk.Label(self.login_frame, text='Password').grid(column=0, row=1)
        ttk.Entry(self.login_frame, textvariable=self.login_pass).grid(
            column=1, row=1, sticky='WE')
        ttk.Button(self.login_frame, text='Login', command=self.login).grid(
            column=1, row=2, sticky='WEN')
        # ---- sign-up
        self.sign_up_frame.grid(column=2, row=5, sticky='NEWS', columnspan=2,
                                rowspan=3, pady=5, padx=5)
        ttk.Label(self.sign_up_frame, text='Name').grid(column=0, row=0)
        ttk.Entry(self.sign_up_frame, textvariable=self.sign_name).grid(
            column=1, row=0, sticky='WE')
        ttk.Label(self.sign_up_frame, text='Password').grid(column=0, row=1)
        ttk.Entry(self.sign_up_frame, textvariable=self.sign_pass).grid(
            column=1, row=1, sticky='WE')
        ttk.Button(self.sign_up_frame, text='Sign-up',
                   command=self.sign_up).grid(
            column=1, row=2, sticky='WEN', columnspan=2)
        # ---- remove
        self.remove_frame.grid(column=2, row=8, sticky='NEWS', columnspan=2,
                               rowspan=2, pady=5, padx=5)
        ttk.Label(self.remove_frame, text='Admin name').grid(
            column=0, row=0)
        ttk.Entry(self.remove_frame, textvariable=self.remove_name).grid(
            column=1, row=0, sticky='WE')
        ttk.Button(self.remove_frame, text='Remove Admin',
                   command=self.remove_admin).grid(column=1, row=2, sticky='WEN')
        # ---- others
        self.other_frame.grid(column=0, row=8, sticky='NEWS', columnspan=2,
                              rowspan=2, pady=5, padx=5)
        ttk.Button(self.other_frame, text='Make Report',
                   command=self.generate_report).grid(
            column=0, row=0, rowspan=2, sticky='NEWS')
        ttk.Button(self.other_frame, text='Lock Access',
                   command=self.lock_access).grid(
            column=1, row=0, sticky='NEWS')
        ttk.Button(self.other_frame, text='Wipe Data',
                   command=self.wipe_data).grid(column=1, row=1, sticky='NEWS')
        # ---- admins
        self.admins_frame.grid(column=0, row=2, columnspan=2, rowspan=6,
                               sticky='NEWS', pady=5, padx=5)
        self.admins_listbox.grid(column=0, row=0, sticky='NEWS', rowspan=6)
        self.admins_s1.grid(column=1, row=0, sticky='NSW', rowspan=6)
        self.admins_s2.grid(column=0, row=6, sticky='WEN')
        # status bar
        ttk.Label(self.master, textvariable=self.status, relief='sunken').grid(
            column=0, row=1, sticky='NEWS')


if __name__ == '__main__':
    root = Tk()
    app = RoomFinanceManager(root, padding=5)
    app.grid(column=0, row=0, sticky='NEWS')
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    root.title('Room Finance Manager')
    root.wm_iconbitmap('icon.ico')
    root.mainloop()
